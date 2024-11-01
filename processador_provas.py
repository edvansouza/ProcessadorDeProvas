import streamlit as st
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import io
import base64

st.set_page_config(page_title="Processador de Provas IBE", page_icon="📊", layout="wide")

def formatar_data(data):
    data_obj = pd.to_datetime(data)
    dias_semana = ['SEG', 'TER', 'QUA', 'QUI', 'SEX', 'SÁB', 'DOM']
    meses = ['JAN', 'FEV', 'MAR', 'ABR', 'MAI', 'JUN', 'JUL', 'AGO', 'SET', 'OUT', 'NOV', 'DEZ']
    dia_semana = dias_semana[data_obj.weekday()]
    dia = data_obj.strftime('%d')
    mes = meses[data_obj.month - 1]
    ano = data_obj.strftime('%y')
    hora_minuto = data_obj.strftime('%H:%M')
    return f"{dia_semana}, {dia}/{mes}/{ano}, {hora_minuto}"

def converter_pontuacao(pontuacao):
    if isinstance(pontuacao, str):
        return 1 if pontuacao.startswith('1') else 0
    elif isinstance(pontuacao, (int, float)):
        return 1 if pontuacao >= 1 else 0
    else:
        return 0

def determinar_numero_de_questoes(df):
    num_colunas = len(df.columns)
    if num_colunas == 156:
        return 5
    elif num_colunas == 96:
        return 3
    else:
        raise ValueError("Número de colunas não corresponde a uma prova de 3 ou 5 questões.")

def processar_arquivo(df):
    # Criar um buffer em memória para o arquivo Excel
    output = io.BytesIO()
    
    # Criar um novo Workbook
    wb = Workbook()
    
    # Determinar número de questões
    num_questoes = determinar_numero_de_questoes(df)
    num_materias = 10
    colunas_por_questao = 3
    
    # Processar cada matéria
    for materia_index in range(num_materias):
        # Criar nova planilha
        start_question = materia_index * num_questoes + 1
        end_question = start_question + num_questoes - 1
        sheet_name = f"Questões {start_question} a {end_question}"
        ws = wb.create_sheet(title=sheet_name)
        
        # Adicionar cabeçalhos
        headers = ['DATA RESPOSTA', 'EMAIL RESPOSTA', 'NOME ALUNO'] + \
                 [f'Q{i+1}' for i in range(num_questoes)] + ['TOTAL']
        ws.append(headers)
        
        # Formatar cabeçalhos
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            cell.font = Font(color="FFFFFF")
        
        # Coletar e ordenar dados
        dados_planilha = []
        for _, row in df.iterrows():
            data_resposta = formatar_data(row['Carimbo de data/hora'])
            email_resposta = row['Nome de usuário']
            nome_aluno = row.iloc[3].upper()
            
            questoes_pontos = []
            start_coluna = 6 + (materia_index * num_questoes * colunas_por_questao)
            for q in range(num_questoes):
                coluna_pontuacao = start_coluna + (q * colunas_por_questao) + 1
                pontos = converter_pontuacao(row[df.columns[coluna_pontuacao]])
                questoes_pontos.append(pontos)
            
            total_pontos = sum(questoes_pontos)
            dados_planilha.append([data_resposta, email_resposta, nome_aluno] + 
                                questoes_pontos + [total_pontos])
        
        # Ordenar por nome do aluno
        dados_ordenados = sorted(dados_planilha, key=lambda x: x[2])
        for linha in dados_ordenados:
            ws.append(linha)
        
        # Formatar células
        for row in ws.iter_rows(min_row=2):
            fill = PatternFill(start_color="CCCCCC" if row[0].row % 2 == 0 else "FFFFFF",
                             end_color="CCCCCC" if row[0].row % 2 == 0 else "FFFFFF",
                             fill_type="solid")
            for cell in row:
                cell.fill = fill
        
        # Formatar totais
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, 
                              min_col=ws.max_column, max_col=ws.max_column):
            for cell in row:
                cell.font = Font(bold=True, size=14)
                if cell.value is not None:
                    if cell.value < 3:
                        cell.fill = PatternFill(start_color="FF9999", fill_type="solid")
                    elif cell.value == 3:
                        cell.fill = PatternFill(start_color="FFFF99", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="CCFFCC", fill_type="solid")
        
        # Ajustar largura das colunas
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2
        
        # Ajustar zoom
        ws.sheet_view.zoomScale = 150
    
    # Remover planilha padrão
    wb.remove(wb['Sheet'])
    
    # Salvar no buffer
    wb.save(output)
    return output

def main():
    st.title("📊 Processador de Provas IBE")
    
    st.markdown("""
    ### Instruções
    1. Faça o upload do arquivo CSV da prova
    2. O sistema processará automaticamente o arquivo
    3. Um arquivo Excel será gerado com as notas organizadas
    """)
    
    uploaded_file = st.file_uploader("Escolha o arquivo CSV da prova", type=['csv'])
    
    if uploaded_file is not None:
        try:
            with st.spinner('Processando o arquivo...'):
                # Ler o arquivo CSV
                df = pd.read_csv(uploaded_file)
                
                # Processar o arquivo
                excel_buffer = processar_arquivo(df)
                
                # Preparar o download
                b64 = base64.b64encode(excel_buffer.getvalue()).decode()
                filename = uploaded_file.name.replace('.csv', '_processado.xlsx')
                
                # Criar botão de download
                st.success('Arquivo processado com sucesso!')
                st.markdown(
                    f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" '
                    f'download="{filename}" class="button">📥 Clique aqui para baixar o arquivo Excel</a>',
                    unsafe_allow_html=True
                )
                
        except Exception as e:
            st.error(f'Erro ao processar o arquivo: {str(e)}')

if __name__ == "__main__":
    main()

# Adicionar estilo CSS personalizado
st.markdown("""
<style>
    .button {
        background-color: #212121;
        border: none;
        color: white;
        padding: 15px 32px;
        text-align: center;
        text-decoration: none;
        display: inline-block;
        font-size: 16px;
        margin: 4px 2px;
        cursor: pointer;
        border-radius: 4px;
    }
    .button:hover {
        background-color: #333333;
    }
</style>
""", unsafe_allow_html=True)
